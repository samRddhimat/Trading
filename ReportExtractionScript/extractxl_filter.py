import csv
from openpyxl import load_workbook

def extract_deal_section_to_csv(excel_file, output_csv, keywords):
    wb = load_workbook(excel_file, read_only=True)
    ws = wb.active

    start_found = False
    rows_to_write = []
    rows = iter(ws.iter_rows(values_only=True))

    for row in rows:
        row_text = [str(cell).strip().lower() if cell is not None else '' for cell in row]
        
        if not start_found:
            if any("deal" in cell for cell in row_text):
                start_found = True
                rows_to_write.append(next(rows))  # include the 'Deal' row
        else:
            if any("open positions" in cell for cell in row_text):
                break  # stop at 'Open Positions' row
            if any(keyword.lower() in cell for cell in row_text for keyword in keywords):
                rows_to_write.append(row)

    # Write to CSV
    with open(output_csv, "w", newline="", encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerows(rows_to_write)

# Usage
keywords_to_match = ["rnd","sfe"] #["deal", "open positions", "usd"]
extract_deal_section_to_csv("D:\\Repository\\Trading\\ReportHistory-205350376.xlsx", \
                            "D:\\Repository\\Trading\\deal_section.csv", \
                                 keywords_to_match)
